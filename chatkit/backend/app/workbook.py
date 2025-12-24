from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Template mapping (must match templates/README.md)
HEADER_YEAR_CELL = "B2"
HEADER_ADDRESS_CELL = "B4"
HEADER_TENANT_CELL = "H4"
HEADER_PERIOD_CELL = "H5"
TABLE_START_ROW = 9

# Columns A..I
COL_MONTH = 1
COL_RENT_DUE = 2
COL_HOUSING_DEPT = 3
COL_HOUSING_PAID = 4
COL_TENANT_PAID = 5
COL_TOTAL_RECEIVED = 6
COL_MONTH_BAL = 7
COL_YEAR_BAL = 8
COL_REMARKS = 9


def _repo_root() -> Path:
    # backend/app/workbook.py -> repo root is: app -> backend -> chatkit
    return Path(__file__).resolve().parents[2]


def generate_rent_workbook(payload: dict[str, Any]) -> bytes:
    template_version = payload.get("template_version")
    if template_version != "property_rents_received_v1":
        raise ValueError("Unsupported template_version")

    year = payload.get("year")
    if year != 2025:
        raise ValueError("Unsupported year")

    template_path = _repo_root() / "templates" / "Property_Rents_Received_Template.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)

    properties = payload.get("properties") or []
    for p in properties:
        property_id = p.get("property_id")
        if not property_id:
            continue

        if property_id not in wb.sheetnames:
            continue

        ws = wb[property_id]

        ws[HEADER_YEAR_CELL].value = year
        ws[HEADER_ADDRESS_CELL].value = p.get("property_address") or ""
        ws[HEADER_TENANT_CELL].value = p.get("tenant_name") or ""
        period_months = p.get("period_months")
        ws[HEADER_PERIOD_CELL].value = period_months if period_months is not None else ""

        rows = p.get("rows") or []
        by_month: dict[int, dict[str, Any]] = {}
        for r in rows:
            try:
                mn = int(r.get("month_number"))
            except Exception:
                continue
            by_month[mn] = r

        for mn in range(1, 13):
            r = by_month.get(mn, {})
            excel_row = TABLE_START_ROW + (mn - 1)

            ws.cell(row=excel_row, column=COL_MONTH).value = r.get("month") or ""
            ws.cell(row=excel_row, column=COL_RENT_DUE).value = r.get("rent_due") or 0
            ws.cell(row=excel_row, column=COL_HOUSING_DEPT).value = r.get("housing_dept") or ""
            ws.cell(row=excel_row, column=COL_HOUSING_PAID).value = r.get("housing_paid") or 0
            ws.cell(row=excel_row, column=COL_TENANT_PAID).value = r.get("tenant_paid") or 0
            ws.cell(row=excel_row, column=COL_TOTAL_RECEIVED).value = r.get("total_received") or 0
            ws.cell(row=excel_row, column=COL_MONTH_BAL).value = r.get("month_balance_due") or 0
            ws.cell(row=excel_row, column=COL_YEAR_BAL).value = r.get("year_balance_due") or 0
            ws.cell(row=excel_row, column=COL_REMARKS).value = r.get("remarks") or ""

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
